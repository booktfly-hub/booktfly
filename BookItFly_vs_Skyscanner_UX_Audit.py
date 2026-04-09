import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# Colors
PRIMARY = "1A56DB"
DARK = "1E293B"
WHITE = "FFFFFF"
LIGHT_BLUE = "EBF5FF"
LIGHT_GREEN = "ECFDF5"
LIGHT_RED = "FEF2F2"
LIGHT_YELLOW = "FFFBEB"
LIGHT_GRAY = "F8FAFC"
GREEN = "059669"
RED = "DC2626"
AMBER = "D97706"
GRAY = "64748B"

header_font = Font(name="Arial", bold=True, color=WHITE, size=11)
header_fill = PatternFill("solid", fgColor=DARK)
subheader_font = Font(name="Arial", bold=True, color=WHITE, size=10)
subheader_fill = PatternFill("solid", fgColor=PRIMARY)
category_font = Font(name="Arial", bold=True, color=DARK, size=10)
category_fill = PatternFill("solid", fgColor="E2E8F0")
body_font = Font(name="Arial", size=10, color="334155")
body_font_bold = Font(name="Arial", size=10, color="334155", bold=True)
wrap = Alignment(wrap_text=True, vertical="top")
wrap_center = Alignment(wrap_text=True, vertical="center", horizontal="center")
thin_border = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)

status_fills = {
    "Missing": PatternFill("solid", fgColor=LIGHT_RED),
    "Needs Update": PatternFill("solid", fgColor=LIGHT_YELLOW),
    "Good": PatternFill("solid", fgColor=LIGHT_GREEN),
    "Partial": PatternFill("solid", fgColor=LIGHT_YELLOW),
}
status_fonts = {
    "Missing": Font(name="Arial", size=10, bold=True, color=RED),
    "Needs Update": Font(name="Arial", size=10, bold=True, color=AMBER),
    "Good": Font(name="Arial", size=10, bold=True, color=GREEN),
    "Partial": Font(name="Arial", size=10, bold=True, color=AMBER),
}
priority_fills = {
    "Critical": PatternFill("solid", fgColor="FEE2E2"),
    "High": PatternFill("solid", fgColor="FEF3C7"),
    "Medium": PatternFill("solid", fgColor="E0F2FE"),
    "Low": PatternFill("solid", fgColor="F1F5F9"),
}
priority_fonts = {
    "Critical": Font(name="Arial", size=10, bold=True, color=RED),
    "High": Font(name="Arial", size=10, bold=True, color=AMBER),
    "Medium": Font(name="Arial", size=10, bold=True, color="0284C7"),
    "Low": Font(name="Arial", size=10, bold=True, color=GRAY),
}

def style_cell(ws, row, col, value, font=body_font, fill=None, alignment=wrap, border=thin_border):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = alignment
    cell.border = border
    return cell

# =====================================================
# SHEET 1: EXECUTIVE SUMMARY
# =====================================================
ws1 = wb.active
ws1.title = "Executive Summary"
ws1.sheet_properties.tabColor = PRIMARY

ws1.column_dimensions["A"].width = 4
ws1.column_dimensions["B"].width = 30
ws1.column_dimensions["C"].width = 18
ws1.column_dimensions["D"].width = 18
ws1.column_dimensions["E"].width = 18
ws1.column_dimensions["F"].width = 18

# Title
ws1.merge_cells("A1:F1")
title_cell = ws1["A1"]
title_cell.value = "BookItFly vs Skyscanner — UX Audit Report"
title_cell.font = Font(name="Arial", bold=True, color=WHITE, size=16)
title_cell.fill = PatternFill("solid", fgColor=DARK)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 45

ws1.merge_cells("A2:F2")
ws1["A2"].value = "Date: April 8, 2026  |  Prepared for: BookItFly Team"
ws1["A2"].font = Font(name="Arial", size=10, color=GRAY, italic=True)
ws1["A2"].fill = PatternFill("solid", fgColor=LIGHT_GRAY)
ws1["A2"].alignment = Alignment(horizontal="center")

# Summary stats
r = 4
ws1.merge_cells(f"A{r}:F{r}")
ws1[f"A{r}"].value = "AUDIT OVERVIEW"
ws1[f"A{r}"].font = Font(name="Arial", bold=True, color=WHITE, size=12)
ws1[f"A{r}"].fill = subheader_fill
ws1[f"A{r}"].alignment = Alignment(horizontal="center")

r = 5
stats = [
    ("", "Category", "Critical", "High", "Medium", "Low"),
]
for col_idx, val in enumerate(stats[0], 1):
    style_cell(ws1, r, col_idx, val, font=Font(name="Arial", bold=True, color=DARK, size=10), fill=PatternFill("solid", fgColor="E2E8F0"), alignment=wrap_center)

categories_summary = [
    ("1", "Search Experience", 2, 3, 2, 0),
    ("2", "Results & Filtering", 1, 4, 2, 1),
    ("3", "Booking Flow", 1, 2, 3, 0),
    ("4", "Trust & Credibility", 1, 3, 2, 0),
    ("5", "Navigation & IA", 0, 2, 3, 1),
    ("6", "Mobile UX", 0, 3, 2, 1),
    ("7", "Accessibility", 1, 2, 2, 1),
    ("8", "Visual Design", 0, 2, 3, 1),
    ("9", "User Accounts", 1, 2, 1, 1),
    ("10", "Performance & States", 0, 2, 3, 0),
]
for i, (num, cat, crit, high, med, low) in enumerate(categories_summary):
    r = 6 + i
    bg = PatternFill("solid", fgColor=LIGHT_BLUE) if i % 2 == 0 else None
    style_cell(ws1, r, 1, num, alignment=wrap_center, fill=bg)
    style_cell(ws1, r, 2, cat, font=body_font_bold, fill=bg)
    c = style_cell(ws1, r, 3, crit, alignment=wrap_center, fill=bg)
    if crit > 0: c.font = Font(name="Arial", size=10, bold=True, color=RED)
    c = style_cell(ws1, r, 4, high, alignment=wrap_center, fill=bg)
    if high > 0: c.font = Font(name="Arial", size=10, bold=True, color=AMBER)
    style_cell(ws1, r, 5, med, alignment=wrap_center, fill=bg)
    style_cell(ws1, r, 6, low, alignment=wrap_center, fill=bg)

r = 16
style_cell(ws1, r, 1, "", font=body_font_bold, fill=PatternFill("solid", fgColor="E2E8F0"))
style_cell(ws1, r, 2, "TOTALS", font=Font(name="Arial", bold=True, color=DARK, size=11), fill=PatternFill("solid", fgColor="E2E8F0"), alignment=wrap_center)
style_cell(ws1, r, 3, "=SUM(C6:C15)", font=Font(name="Arial", bold=True, color=RED, size=11), fill=PatternFill("solid", fgColor="FEE2E2"), alignment=wrap_center)
style_cell(ws1, r, 4, "=SUM(D6:D15)", font=Font(name="Arial", bold=True, color=AMBER, size=11), fill=PatternFill("solid", fgColor="FEF3C7"), alignment=wrap_center)
style_cell(ws1, r, 5, "=SUM(E6:E15)", font=Font(name="Arial", bold=True, color="0284C7", size=11), fill=PatternFill("solid", fgColor="E0F2FE"), alignment=wrap_center)
style_cell(ws1, r, 6, "=SUM(F6:F15)", font=Font(name="Arial", bold=True, color=GRAY, size=11), fill=PatternFill("solid", fgColor="F1F5F9"), alignment=wrap_center)

# Key findings
r = 18
ws1.merge_cells(f"A{r}:F{r}")
ws1[f"A{r}"].value = "TOP 7 CRITICAL & HIGH-PRIORITY GAPS"
ws1[f"A{r}"].font = Font(name="Arial", bold=True, color=WHITE, size=12)
ws1[f"A{r}"].fill = PatternFill("solid", fgColor=RED)
ws1[f"A{r}"].alignment = Alignment(horizontal="center")

key_gaps = [
    ("1", "No flexible date search / price calendar", "Skyscanner's price calendar lets users find cheapest dates across a month. BookItFly only has single date pickers.", "Critical"),
    ("2", "No multi-city / nearby airports search", "Skyscanner supports multi-city itineraries and nearby airport suggestions. BookItFly only supports one-way/round-trip.", "Critical"),
    ("3", "No price alerts / saved searches", "Skyscanner lets users set price alerts and save searches. BookItFly has no equivalent — users must manually recheck.", "Critical"),
    ("4", "Weak filtering on results page", "Skyscanner has duration slider, stops filter, time range, airline filter. BookItFly only has price range and cabin class.", "High"),
    ("5", "No 'Best' sort algorithm", "Skyscanner's default sort weighs price + duration + stops. BookItFly only sorts by newest, price, or date.", "High"),
    ("6", "No price comparison across providers", "Skyscanner shows the same route from multiple providers with price comparison. BookItFly shows one provider per listing.", "High"),
    ("7", "Missing social proof on results", "Skyscanner shows ratings, review counts, and 'popular choice' badges. BookItFly only shows seat count and provider name.", "High"),
]

r = 19
headers = ("", "Gap", "Details", "Priority")
cols = [1, 2, 3, 6]
for i, (ci, val) in enumerate(zip(cols, headers)):
    if i == 2:
        ws1.merge_cells(f"C{r}:E{r}")
    if i == 3:
        pass  # merged
    style_cell(ws1, r, ci, val, font=Font(name="Arial", bold=True, color=DARK, size=10), fill=PatternFill("solid", fgColor="E2E8F0"), alignment=wrap_center)

for i, (num, gap, detail, priority) in enumerate(key_gaps):
    r = 20 + i
    ws1.row_dimensions[r].height = 45
    ws1.merge_cells(f"C{r}:E{r}")
    style_cell(ws1, r, 1, num, alignment=wrap_center)
    style_cell(ws1, r, 2, gap, font=body_font_bold)
    style_cell(ws1, r, 3, detail)
    c = style_cell(ws1, r, 6, priority, alignment=wrap_center)
    c.font = priority_fonts.get(priority, body_font)
    c.fill = priority_fills.get(priority, None)


# =====================================================
# SHEET 2: DETAILED COMPARISON
# =====================================================
ws2 = wb.create_sheet("Detailed Comparison")
ws2.sheet_properties.tabColor = "059669"

cols_width = {"A": 4, "B": 22, "C": 35, "D": 35, "E": 14, "F": 12, "G": 40}
for col, w in cols_width.items():
    ws2.column_dimensions[col].width = w

# Title
ws2.merge_cells("A1:G1")
ws2["A1"].value = "BookItFly vs Skyscanner — Feature-by-Feature Comparison"
ws2["A1"].font = Font(name="Arial", bold=True, color=WHITE, size=14)
ws2["A1"].fill = PatternFill("solid", fgColor=DARK)
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 40

# Column headers
r = 2
col_headers = ["#", "Feature", "Skyscanner Has", "BookItFly Has", "Status", "Priority", "Recommended Action"]
for ci, h in enumerate(col_headers, 1):
    style_cell(ws2, r, ci, h, font=header_font, fill=header_fill, alignment=wrap_center)
ws2.row_dimensions[2].height = 30

# Data
comparison_data = [
    # SEARCH EXPERIENCE
    ("SEARCH EXPERIENCE", None, None, None, None, None, None),
    ("1", "Flexible date search / price calendar", "Monthly view showing cheapest dates with color-coded prices across the entire month", "Single date picker only — no price visibility for nearby dates", "Missing", "Critical", "Add a 'Flexible Dates' toggle next to the date picker that shows a monthly calendar with prices overlaid on each date"),
    ("2", "Multi-city search", "Supports adding multiple flight legs (city A→B→C→D) in a single search", "Only One Way and Round Trip options available", "Missing", "Critical", "Add a 'Multi-city' tab to the trip type selector with dynamic leg inputs"),
    ("3", "Nearby airports suggestion", "Suggests nearby airports automatically (e.g., searching London shows LHR, LGW, STN, LTN)", "Single airport per city via autocomplete — no nearby alternatives shown", "Missing", "High", "Enhance city-autocomplete to show 'Nearby airports' group when a city has multiple airports"),
    ("4", "Cabin class in search form", "Economy, Premium Economy, Business, First Class selectable in the main search bar", "Cabin class only available as a filter AFTER search results load", "Needs Update", "High", "Move cabin class selector into the hero search form so users set expectations before searching"),
    ("5", "Passenger age categories", "Adults (16+), Children (1-16), Infants (<1) with age inputs per child", "Simple numeric passenger counter with no age differentiation", "Needs Update", "High", "Add passenger type breakdown (Adults / Children / Infants) with age selectors for children"),
    ("6", "Swap origin/destination animation", "Smooth animated swap with icon rotation", "Swap button exists with basic icon rotation", "Good", "Low", "Current implementation is acceptable"),
    ("7", "Search form field reduction", "Reduced from 5 to 3 visible fields to minimize cognitive load — advanced options hidden behind toggles", "All fields visible at once (origin, dest, trip type, depart, return, passengers)", "Needs Update", "Medium", "Consider progressive disclosure: show only origin + destination + dates initially, expand for more options"),

    # RESULTS & FILTERING
    ("RESULTS & FILTERING", None, None, None, None, None, None),
    ("8", "'Best' sort algorithm", "Default sort combines price + duration + stops into a weighted 'Best' score", "Sort options: newest, price asc/desc, date — no composite quality score", "Missing", "High", "Implement a 'Best' sorting algorithm that factors in price, duration, stops, and provider rating"),
    ("9", "Flight duration filter", "Slider to filter by maximum flight duration", "No duration-based filtering available", "Missing", "High", "Add a range slider for maximum flight duration in the filter panel"),
    ("10", "Number of stops filter", "Filter by Direct, 1 Stop, 2+ Stops", "No stops/layover filtering", "Missing", "High", "Add a stops filter with options: Direct, 1 Stop, 2+ Stops (if your data model supports layovers)"),
    ("11", "Departure/arrival time filter", "Time range sliders for departure and arrival windows (e.g., morning, afternoon, evening)", "No time-based filtering", "Missing", "Medium", "Add time-of-day filter with preset ranges (Morning 6-12, Afternoon 12-18, Evening 18-24)"),
    ("12", "Airline/provider filter", "Checkbox list of all airlines serving the route, with logos", "No provider filtering in results — all providers mixed", "Missing", "High", "Add a provider/airline filter sidebar section with checkboxes and provider logos"),
    ("13", "Price per route comparison", "Same route shows prices from multiple providers side by side for easy comparison", "Each trip is a separate card — no side-by-side price comparison for same route", "Missing", "High", "Group trips by route and show a 'Compare prices' view with all providers for the same origin→destination pair"),
    ("14", "Results count display", "Shows 'X results found' at top of results", "No visible result count — user doesn't know total available options", "Missing", "Medium", "Add a results count badge (e.g., '24 flights found') above the results grid"),
    ("15", "Inline search modification", "Search bar stays visible and editable at top of results — no page reload needed", "Search bar present at top of results page with client-side filtering", "Good", "Low", "Current implementation is solid"),

    # BOOKING FLOW
    ("BOOKING FLOW", None, None, None, None, None, None),
    ("16", "Price breakdown transparency", "Full breakdown: base fare + taxes + fees + service charge shown before checkout", "Shows per-seat price × count = total — no fee breakdown", "Needs Update", "Critical", "Add a detailed price breakdown showing base fare, taxes, service fees, and any platform commission separately"),
    ("17", "Seat selection UX", "Visual aircraft seat map with clear tier pricing and real-time availability", "Interactive seat map with color-coded tiers (economy, extra legroom, up-front) and aria-labels", "Good", "Low", "Seat map is well-implemented — consider adding tooltip with tier pricing on hover"),
    ("18", "Passport/ID scanning", "Not a standard feature", "Has passport scan functionality for auto-filling passenger details", "Good", "Low", "This is actually ahead of Skyscanner — great differentiator"),
    ("19", "Guest checkout", "Supports searching without login; account created during booking", "Has guest booking via token links, but registration required for direct booking", "Needs Update", "High", "Allow users to complete booking without mandatory registration — offer optional account creation post-purchase"),
    ("20", "Progress indicator in booking", "Clear step indicator: Search → Select → Passenger Details → Payment → Confirmation", "No visible step/progress indicator — user doesn't know how many steps remain", "Missing", "High", "Add a horizontal stepper (Step 1: Select, Step 2: Details, Step 3: Payment, Step 4: Confirm) at the top of booking flow"),
    ("21", "Payment methods", "Redirects to airline/provider sites which offer credit card, PayPal, etc.", "Bank transfer only with IBAN and receipt upload", "Needs Update", "Medium", "While bank transfer fits the MENA market, consider adding online payment options (Mada, Apple Pay, credit cards) for convenience"),
    ("22", "Booking confirmation email", "Immediate email with full itinerary, booking ref, and calendar invite", "Email system exists (Resend + React Email), confirmation notifications sent", "Needs Update", "Medium", "Ensure confirmation emails include: full itinerary, booking reference, calendar .ics attachment, and provider contact info"),
    ("23", "Cancellation policy visibility", "Clearly displayed cancellation terms and refund policy before checkout", "No visible cancellation policy on booking or checkout pages", "Missing", "Medium", "Add a collapsible 'Cancellation Policy' section on trip detail and checkout pages"),

    # TRUST & CREDIBILITY
    ("TRUST & CREDIBILITY", None, None, None, None, None, None),
    ("24", "User reviews & ratings", "Star ratings and review counts displayed on results and detail pages", "FlyPoints system exists but no public user reviews or ratings on trips", "Missing", "Critical", "Implement a review/rating system where users can rate providers and trips after travel — display aggregated ratings on cards"),
    ("25", "Provider verification badges", "Partner airline logos and trust indicators", "Provider application system with document verification, but no visible trust badge on public pages", "Needs Update", "High", "Add a 'Verified Provider' badge on trip cards and detail pages for approved providers"),
    ("26", "'Popular choice' / social proof badges", "Badges like 'Most popular', 'Best value', 'Cheapest' on top results", "Last-minute badge and urgency messaging ('Only X left!') exist", "Partial", "High", "Add algorithmic badges: 'Most Booked', 'Best Value', 'Top Rated' based on actual booking and rating data"),
    ("27", "Security badges on checkout", "SSL, secure payment icons, and partner logos on payment page", "Bank transfer instructions shown but no security indicators on checkout page", "Missing", "High", "Add security trust signals: SSL badge, secure transaction messaging, and payment method logos on the checkout page"),
    ("28", "Help center / FAQ", "Comprehensive help.skyscanner.net with searchable FAQ, categories, contact options", "No visible help center, FAQ, or support chat", "Missing", "Medium", "Add a help/FAQ page with common questions about booking, cancellation, payment, and provider policies"),
    ("29", "Company 'About' page", "About page with company history, mission, investor info", "No about page — footer has mission statement only", "Missing", "Medium", "Create an About page with company story, team, mission, and press/media mentions to build credibility"),

    # NAVIGATION & INFORMATION ARCHITECTURE
    ("NAVIGATION & INFORMATION ARCHITECTURE", None, None, None, None, None, None),
    ("30", "Breadcrumb navigation", "Breadcrumbs on deeper pages for orientation (Home > Flights > London > Results)", "No breadcrumb navigation on any page", "Missing", "Medium", "Add breadcrumbs on trip detail, booking, and checkout pages to help users orient themselves"),
    ("31", "Category tabs in search", "Clear tabs for Flights, Hotels, Car Hire at top of search", "Tabs for Flights, Hotels, Cars in hero section", "Good", "Low", "Well implemented"),
    ("32", "Sticky search/filter bar", "Search bar stays fixed on scroll in results page for quick modifications", "Search bar at top of results but not sticky", "Needs Update", "High", "Make the search bar sticky on the results page so users can modify search without scrolling to top"),
    ("33", "Footer navigation structure", "Organized footer with Explore, Company, Partners, Help sections plus legal links", "Footer has Discover, Company, Legal sections with social links and trust badges", "Good", "Low", "Footer is solid — consider adding a Partner/Provider section for B2B visibility"),
    ("34", "Back navigation consistency", "Browser back button works correctly throughout the booking flow", "Back button on trip detail page — needs testing across full flow", "Needs Update", "Medium", "Ensure consistent back navigation using Next.js router throughout all pages, especially in booking flow"),
    ("35", "Search URL state preservation", "Search parameters encoded in URL for sharing and browser history", "URL parameters used for search state — params encoded in query string", "Good", "Low", "Well implemented"),

    # MOBILE UX
    ("MOBILE UX", None, None, None, None, None, None),
    ("36", "Bottom tab navigation (mobile)", "Bottom tab bar with 4 sections for quick mobile navigation", "Hamburger menu only — no bottom tab navigation on mobile", "Missing", "High", "Add a fixed bottom tab bar on mobile with: Home, Search, My Bookings, Profile icons"),
    ("37", "Mobile search form optimization", "Reduced input fields with popover menus to minimize keyboard switching", "Responsive grid layout adapts, but all fields shown at once", "Needs Update", "High", "Redesign mobile search to use full-screen modals for each input (tap Origin → full-screen city selector with recent searches)"),
    ("38", "Touch-friendly filter UI", "Swipeable filter chips and bottom-sheet filter panels optimized for thumbs", "Filter panel with standard inputs — no bottom-sheet or chip-based interaction", "Needs Update", "High", "Replace the filter panel with a bottom-sheet modal on mobile with large touch targets and swipeable chips"),
    ("39", "Pull-to-refresh", "Standard mobile pattern for refreshing results", "No pull-to-refresh implemented", "Missing", "Medium", "Add pull-to-refresh on results pages for mobile users"),
    ("40", "Mobile sticky booking bar", "Price and CTA button fixed at bottom of screen on detail pages", "Mobile sticky bottom bar implemented with price + book button", "Good", "Low", "Well implemented"),
    ("41", "Swipeable image galleries", "Touch-swipeable photo galleries for hotels and destinations", "Room image gallery exists — need to verify swipe support", "Needs Update", "Medium", "Ensure room and car image galleries support native touch swipe gestures"),

    # ACCESSIBILITY
    ("ACCESSIBILITY", None, None, None, None, None, None),
    ("42", "WCAG 2.2 AA compliance", "Targets WCAG 2.2 Level AA with dedicated accessibility team", "Some aria-labels (seat buttons), proper input types, but no systematic WCAG audit", "Needs Update", "Critical", "Conduct a full WCAG 2.2 AA audit and address: focus indicators, color contrast, screen reader labels, keyboard navigation"),
    ("43", "Keyboard navigation", "Full keyboard navigation support for all interactive elements", "Calendar popovers have keyboard support, but no visible focus indicators on most elements", "Needs Update", "High", "Add visible focus ring styles (:focus-visible) to all interactive elements: buttons, links, cards, inputs"),
    ("44", "Screen reader support", "Alt text, ARIA labels, and landmark roles throughout", "Limited aria-labels (mainly seat buttons), missing on many interactive elements", "Needs Update", "High", "Add aria-labels to: nav items, filter controls, sort buttons, card actions, icon-only buttons, status badges"),
    ("45", "Color contrast compliance", "Working toward AA contrast ratios with documented color palette", "Custom color palette with slate-based grays — contrast not verified", "Needs Update", "Medium", "Run automated contrast checks on all text/background combinations and fix any ratios below 4.5:1"),
    ("46", "Skip navigation link", "Skip-to-content links for keyboard users", "No skip navigation link", "Missing", "Medium", "Add a 'Skip to main content' link that appears on tab focus at the top of every page"),
    ("47", "Error announcement for screen readers", "Live regions for dynamic error messages", "Error messages displayed visually but no aria-live regions", "Missing", "Low", "Add aria-live='polite' regions for form validation errors and search result updates"),

    # VISUAL DESIGN
    ("VISUAL DESIGN", None, None, None, None, None, None),
    ("48", "Consistent design system", "Backpack design system with documented tokens, components across all platforms", "Tailwind CSS + shadcn/ui — consistent but not formally documented as a system", "Needs Update", "Medium", "Document your design tokens (colors, spacing, typography) in a shared reference — consider Storybook for component library"),
    ("49", "Illustration / empty state art", "Custom illustrations for empty states, errors, and onboarding", "EmptyState component with Lucide icons — no custom illustrations", "Needs Update", "High", "Commission or create custom SVG illustrations for: empty search results, no bookings, error pages, onboarding steps"),
    ("50", "Loading skeleton quality", "Content-aware skeletons that match the exact layout of the loaded content", "CardSkeleton and BookingPageSkeleton with animate-pulse", "Good", "Low", "Skeletons are well-matched to content layout — good implementation"),
    ("51", "Micro-animations on interactions", "Subtle transitions on button press, card hover, page transitions", "animate-fade-in-up, hover scale/shadow, active:scale-[0.98] on buttons", "Good", "Low", "Good animation foundation — consider adding page transition animations with Framer Motion"),
    ("52", "Dark mode support", "Not a core feature on Skyscanner", "No dark mode", "Missing", "Medium", "Consider adding dark mode toggle — increasingly expected by users, especially in MENA region (late-night browsing)"),
    ("53", "Consistent iconography", "Unified icon set from Backpack design system", "Lucide React icons used consistently throughout", "Good", "Low", "Good consistency with Lucide"),

    # USER ACCOUNTS
    ("USER ACCOUNTS", None, None, None, None, None, None),
    ("54", "Price alerts", "Email alerts when prices change for saved routes — users can monitor multiple destinations", "No price alert system", "Missing", "Critical", "Build a price alert system: users save routes, receive email/push notifications when prices drop below threshold"),
    ("55", "Saved searches / wishlists", "Save flights and hotels to lists for later comparison and booking", "No saved searches or wishlist functionality", "Missing", "High", "Add a 'Save' heart icon on trip/room/car cards that saves to user's wishlist with quick comparison view"),
    ("56", "Recent searches history", "Shows recent searches on homepage for quick re-search", "localStorage persistence for trip type and passenger count — but no recent search history", "Needs Update", "High", "Store and display last 5 searches on homepage hero section for quick repeat searches"),
    ("57", "Social login options", "Google, Facebook, Apple sign-in", "Google OAuth supported, email/password, magic link", "Needs Update", "Medium", "Add Apple Sign-In and consider Facebook login for broader coverage"),
    ("58", "User profile completeness", "Basic profile with saved payment methods and preferences", "Profile management with avatar, but no saved payment methods or travel preferences", "Needs Update", "Low", "Add travel preferences (preferred cabin class, dietary needs, passport details) to profile for faster booking"),

    # PERFORMANCE & STATES
    ("PERFORMANCE & STATES", None, None, None, None, None, None),
    ("59", "Search results loading UX", "Search bar reduces height, results stream in below without page navigation", "Full page navigation to /trips with loading skeletons", "Needs Update", "High", "Consider rendering results on same page or adding a smooth transition animation between search and results"),
    ("60", "Error state handling", "Graceful error pages with retry options and help links", "EmptyState component for no results, but limited error boundary handling", "Needs Update", "High", "Add proper error boundaries with retry buttons, help links, and fallback UI for API failures"),
    ("61", "Offline / slow connection handling", "Progressive loading with cached content", "No offline handling or service worker", "Missing", "Medium", "Add a service worker for basic offline support and 'No internet' messaging"),
    ("62", "Infinite scroll vs pagination", "Smooth infinite scroll with 'Show more' option", "Load More button with separate loading state", "Needs Update", "Medium", "Consider converting to infinite scroll with intersection observer for smoother browsing on mobile"),
    ("63", "Toast/notification feedback", "Subtle toast notifications for user actions", "Toaster component with toast notifications implemented", "Good", "Low", "Well implemented"),
]

row = 3
item_num = 0
for data in comparison_data:
    if data[1] is None:  # Category header
        ws2.merge_cells(f"A{row}:G{row}")
        ws2[f"A{row}"].value = data[0]
        ws2[f"A{row}"].font = Font(name="Arial", bold=True, color=WHITE, size=11)
        ws2[f"A{row}"].fill = subheader_fill
        ws2[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws2.row_dimensions[row].height = 28
    else:
        ws2.row_dimensions[row].height = 55
        bg = PatternFill("solid", fgColor=LIGHT_GRAY) if item_num % 2 == 0 else None
        style_cell(ws2, row, 1, data[0], alignment=wrap_center, fill=bg)
        style_cell(ws2, row, 2, data[1], font=body_font_bold, fill=bg)
        style_cell(ws2, row, 3, data[2], fill=bg)
        style_cell(ws2, row, 4, data[3], fill=bg)

        status = data[4]
        c = style_cell(ws2, row, 5, status, alignment=wrap_center)
        c.font = status_fonts.get(status, body_font)
        c.fill = status_fills.get(status, bg)

        priority = data[5]
        c = style_cell(ws2, row, 6, priority, alignment=wrap_center)
        c.font = priority_fonts.get(priority, body_font)
        c.fill = priority_fills.get(priority, bg)

        style_cell(ws2, row, 7, data[6], fill=bg)
        item_num += 1
    row += 1

# Freeze panes
ws2.freeze_panes = "A3"


# =====================================================
# SHEET 3: ACTION PLAN (Prioritized)
# =====================================================
ws3 = wb.create_sheet("Action Plan")
ws3.sheet_properties.tabColor = "D97706"

cols3 = {"A": 4, "B": 8, "C": 30, "D": 50, "E": 14, "F": 14, "G": 16}
for col, w in cols3.items():
    ws3.column_dimensions[col].width = w

ws3.merge_cells("A1:G1")
ws3["A1"].value = "Implementation Action Plan — Prioritized Roadmap"
ws3["A1"].font = Font(name="Arial", bold=True, color=WHITE, size=14)
ws3["A1"].fill = PatternFill("solid", fgColor=DARK)
ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 40

r = 2
for ci, h in enumerate(["#", "Phase", "Action Item", "Implementation Notes", "Priority", "Effort", "Impact"], 1):
    style_cell(ws3, r, ci, h, font=header_font, fill=header_fill, alignment=wrap_center)

actions = [
    # Phase 1 - Critical
    ("PHASE 1: CRITICAL FIXES (Weeks 1-4)", None, None, None, None, None),
    ("1", "P1", "Add user reviews & ratings system", "Create reviews table in Supabase, rating component (1-5 stars), display aggregated ratings on trip cards and detail pages. Allow reviews only after completed travel.", "Critical", "Large", "Very High"),
    ("2", "P1", "Build price alerts system", "Add price_alerts table, cron job to check price changes, integrate with Resend email. UI: bell icon on trip cards, 'Set Alert' button on search results, manage alerts in profile.", "Critical", "Large", "Very High"),
    ("3", "P1", "Implement flexible date / price calendar", "Create a calendar overlay component showing lowest prices per day for a given route. Requires API endpoint aggregating trip prices by date range. Toggle on/off in date picker.", "Critical", "Large", "Very High"),
    ("4", "P1", "Add multi-city search support", "Extend trip type from [one-way, round-trip] to include multi-city. Add dynamic leg inputs in search form. Update API to handle multi-leg queries.", "Critical", "Large", "High"),
    ("5", "P1", "Detailed price breakdown on booking", "Break down total price into: base fare, taxes, service fee, platform fee. Show this on trip detail sidebar AND checkout page. Even if fees are zero, showing the breakdown builds trust.", "Critical", "Small", "Very High"),
    ("6", "P1", "WCAG 2.2 AA accessibility audit", "Run axe-core or Lighthouse accessibility audit. Priority fixes: focus indicators, color contrast (4.5:1), aria-labels on all interactive elements, skip navigation link, aria-live for dynamic content.", "Critical", "Medium", "High"),

    # Phase 2 - High Priority
    ("PHASE 2: HIGH-PRIORITY IMPROVEMENTS (Weeks 5-10)", None, None, None, None, None),
    ("7", "P2", "Enhanced filtering system", "Add to filter panel: duration slider, stops filter (Direct/1/2+), time-of-day presets, airline/provider checkboxes. Use URL params so filters are shareable.", "High", "Medium", "Very High"),
    ("8", "P2", "Implement 'Best' sort algorithm", "Create scoring function: weight × (normalized_price + normalized_duration + provider_rating + stops_penalty). Make 'Best' the default sort option.", "High", "Medium", "High"),
    ("9", "P2", "Booking flow progress stepper", "Add a 4-step progress bar: Select → Details → Payment → Confirmed. Show on trip detail, booking form, checkout, and confirmation pages.", "High", "Small", "High"),
    ("10", "P2", "Guest checkout without registration", "Allow booking without signup. Collect email + phone for communication. Offer 'Create account to track your booking' post-purchase.", "High", "Medium", "High"),
    ("11", "P2", "Saved searches & wishlist", "Add saved_items table. Heart icon on cards. 'Saved' page accessible from profile dropdown. Quick comparison grid view.", "High", "Medium", "High"),
    ("12", "P2", "Recent search history", "Store last 5 searches in DB (logged in) or localStorage (guest). Display as horizontal chips below hero search bar on homepage.", "High", "Small", "High"),
    ("13", "P2", "Social proof badges", "Algorithm to tag trips: 'Most Booked' (top 10% by bookings), 'Best Value' (lowest price/quality ratio), 'Top Rated' (4.5+ stars). Display badges on trip cards.", "High", "Medium", "High"),
    ("14", "P2", "Provider verification badges", "Add 'verified' boolean to provider profile. Show blue checkmark badge on trip cards and detail pages for verified providers.", "High", "Small", "High"),
    ("15", "P2", "Security badges on checkout", "Add trust section above payment: SSL padlock icon, 'Secure Transaction' text, accepted payment method logos, money-back guarantee messaging.", "High", "Small", "Very High"),
    ("16", "P2", "Sticky search bar on results", "Make the search/filter bar position:sticky on the trips results page so users can modify search without scrolling.", "High", "Small", "High"),
    ("17", "P2", "Mobile bottom tab navigation", "Add fixed bottom nav bar (visible <768px): Home, Search, Bookings, Profile. Use Next.js usePathname for active state.", "High", "Medium", "High"),
    ("18", "P2", "Mobile search full-screen modals", "On mobile, tapping each search field opens a full-screen modal (city selector with recent + popular, date picker, passengers). Reduces cognitive load.", "High", "Medium", "High"),
    ("19", "P2", "Mobile bottom-sheet filters", "Replace filter panel with a bottom-sheet (slide up from bottom) on mobile. Use large touch targets and filter chips.", "High", "Medium", "High"),
    ("20", "P2", "Custom empty state illustrations", "Create or source SVG illustrations for: no results, no bookings, error states, onboarding. Replace generic Lucide icons.", "High", "Medium", "Medium"),
    ("21", "P2", "Error boundaries with retry", "Wrap major page sections in React error boundaries. Show friendly error UI with 'Try Again' button and link to help/support.", "High", "Small", "High"),
    ("22", "P2", "Visible focus indicators", "Add :focus-visible ring styles to all buttons, links, inputs, cards. Use ring-2 ring-primary ring-offset-2 pattern from Tailwind.", "High", "Small", "High"),
    ("23", "P2", "Screen reader aria-labels", "Audit all interactive elements. Add aria-labels to: icon-only buttons, status badges, nav items, card actions, filter controls, sort dropdowns.", "High", "Small", "High"),

    # Phase 3 - Medium Priority
    ("PHASE 3: ENHANCEMENTS (Weeks 11-16)", None, None, None, None, None),
    ("24", "P3", "Nearby airports suggestion", "Enhance city autocomplete: group results by city and show all airports. Add 'Include nearby airports' toggle.", "Medium", "Medium", "Medium"),
    ("25", "P3", "Cabin class in hero search", "Move cabin class selector from results filter into the hero search form dropdown.", "Medium", "Small", "Medium"),
    ("26", "P3", "Passenger age categories", "Replace simple passenger counter with Adult/Child/Infant breakdown. Child age selectors for each child.", "Medium", "Medium", "Medium"),
    ("27", "P3", "Results count display", "Add 'X flights found' badge above results grid. Update count when filters change.", "Medium", "Small", "Medium"),
    ("28", "P3", "Breadcrumb navigation", "Add breadcrumbs to trip detail, booking, checkout pages. Use Next.js route segments.", "Medium", "Small", "Medium"),
    ("29", "P3", "Help center / FAQ page", "Create /help route with categorized FAQs: Booking, Payment, Cancellation, Provider, Account. Add search functionality.", "Medium", "Medium", "Medium"),
    ("30", "P3", "About page", "Create /about route with company story, team, mission statement, and any press mentions.", "Medium", "Small", "Medium"),
    ("31", "P3", "Cancellation policy display", "Add collapsible policy section on trip detail and checkout. Pull policy text from trip or provider data.", "Medium", "Small", "High"),
    ("32", "P3", "Booking confirmation email enhancement", "Ensure emails include: full itinerary, booking ref, .ics calendar attachment, provider contact, cancellation policy link.", "Medium", "Medium", "Medium"),
    ("33", "P3", "Color contrast fixes", "Run automated contrast check. Fix any text/bg combinations below 4.5:1 AA ratio. Pay attention to gray text on white backgrounds.", "Medium", "Small", "Medium"),
    ("34", "P3", "Skip navigation link", "Add visually hidden 'Skip to main content' link that becomes visible on Tab focus. Place before navbar.", "Medium", "Small", "Medium"),
    ("35", "P3", "Dark mode", "Add dark mode using Tailwind's dark: variant. Store preference in localStorage. Toggle in navbar.", "Medium", "Large", "Medium"),
    ("36", "P3", "Service worker / offline messaging", "Add basic service worker for offline detection. Show 'No internet connection' banner when offline.", "Medium", "Medium", "Low"),
    ("37", "P3", "Infinite scroll on results", "Replace 'Load More' button with intersection observer for automatic loading on scroll.", "Medium", "Small", "Medium"),

    # Phase 4 - Low / Nice-to-have
    ("PHASE 4: POLISH (Weeks 17+)", None, None, None, None, None),
    ("38", "P4", "Progressive search form disclosure", "Default hero search shows 3 fields. 'More options' expands to show all fields. Reduces visual clutter.", "Low", "Medium", "Medium"),
    ("39", "P4", "Design system documentation", "Set up Storybook or similar. Document color tokens, spacing scale, component variants, and usage guidelines.", "Low", "Large", "Medium"),
    ("40", "P4", "Page transition animations", "Add Framer Motion page transitions for smoother navigation between routes.", "Low", "Small", "Low"),
    ("41", "P4", "Apple Sign-In", "Add Apple OAuth provider in Supabase auth for iOS users.", "Low", "Small", "Low"),
    ("42", "P4", "Travel preferences in profile", "Add preferred cabin class, dietary needs, saved passport details to user profile for faster booking.", "Low", "Medium", "Low"),
    ("43", "P4", "Pull-to-refresh on mobile", "Add pull-to-refresh gesture handler for mobile results pages.", "Low", "Small", "Low"),
    ("44", "P4", "Swipeable image galleries", "Ensure room/car image galleries support native touch swipe using a lightweight carousel library.", "Low", "Small", "Low"),
    ("45", "P4", "Aria-live regions for dynamic content", "Add aria-live='polite' to search results count, filter updates, and form validation messages.", "Low", "Small", "Medium"),
]

row = 3
for data in actions:
    if data[1] is None:  # Phase header
        ws3.merge_cells(f"A{row}:G{row}")
        ws3[f"A{row}"].value = data[0]
        ws3[f"A{row}"].font = Font(name="Arial", bold=True, color=WHITE, size=11)
        ws3[f"A{row}"].fill = subheader_fill
        ws3[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws3.row_dimensions[row].height = 28
    else:
        ws3.row_dimensions[row].height = 60
        bg = None
        style_cell(ws3, row, 1, data[0], alignment=wrap_center, fill=bg)
        style_cell(ws3, row, 2, data[1], alignment=wrap_center, font=body_font_bold, fill=bg)
        style_cell(ws3, row, 3, data[2], font=body_font_bold, fill=bg)
        style_cell(ws3, row, 4, data[3], fill=bg)

        priority = data[4]
        c = style_cell(ws3, row, 5, priority, alignment=wrap_center)
        c.font = priority_fonts.get(priority, body_font)
        c.fill = priority_fills.get(priority, bg)

        effort_colors = {"Small": GREEN, "Medium": AMBER, "Large": RED}
        c = style_cell(ws3, row, 6, data[5], alignment=wrap_center)
        c.font = Font(name="Arial", size=10, bold=True, color=effort_colors.get(data[5], GRAY))

        style_cell(ws3, row, 7, data[6], alignment=wrap_center, font=body_font_bold)
    row += 1

ws3.freeze_panes = "A3"


# =====================================================
# SHEET 4: WHAT BOOKITFLY DOES BETTER
# =====================================================
ws4 = wb.create_sheet("BookItFly Strengths")
ws4.sheet_properties.tabColor = "059669"

cols4 = {"A": 4, "B": 28, "C": 55, "D": 45}
for col, w in cols4.items():
    ws4.column_dimensions[col].width = w

ws4.merge_cells("A1:D1")
ws4["A1"].value = "Where BookItFly Is Ahead of Skyscanner"
ws4["A1"].font = Font(name="Arial", bold=True, color=WHITE, size=14)
ws4["A1"].fill = PatternFill("solid", fgColor="059669")
ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 40

r = 2
for ci, h in enumerate(["#", "Feature", "What BookItFly Has", "Why It Matters"], 1):
    style_cell(ws4, r, ci, h, font=header_font, fill=header_fill, alignment=wrap_center)

strengths = [
    ("1", "Passport/ID Scanning", "AI-powered passport scanning that auto-fills passenger details from a photo of the passport document", "Massive time-saver for multi-passenger bookings. Skyscanner doesn't offer this. Highlight it as a key differentiator in marketing."),
    ("2", "Interactive Seat Map", "Visual aircraft seat map with color-coded tiers (economy, extra legroom, up-front), real-time availability, and aria-labels", "Direct seat selection before booking gives users control. Skyscanner redirects to airlines for seat selection."),
    ("3", "Trip Request System", "Users can post flight requests with desired routes/dates, and providers submit competing offers", "Reverse marketplace model — unique in travel. Users get personalized quotes. This is a major competitive advantage."),
    ("4", "Provider Marketplace", "Multi-provider platform where travel agencies/Hajj companies list directly, with admin moderation", "Unlike Skyscanner (aggregator), BookItFly is a marketplace. Direct provider relationships can mean better prices and service."),
    ("5", "Marketeer/Affiliate System", "Third-party agents can book on behalf of customers, track commissions, manage campaigns", "Enables B2B2C distribution — travel agents and influencers can drive bookings with built-in commission tracking."),
    ("6", "FlyPoints Rewards", "Gamification system rewarding bookings, referrals, and reviews with redeemable points", "Loyalty program drives repeat usage. Skyscanner has no equivalent rewards system."),
    ("7", "RTL / Arabic-First Design", "Full Arabic localization with RTL layout support, built into the architecture from day one", "Critical for MENA market. Skyscanner's Arabic support is secondary/translated. BookItFly is culturally native."),
    ("8", "Real-Time Notifications", "Supabase Realtime-powered push notifications for bookings, offers, cancellations", "Instant updates keep users informed. More immediate than Skyscanner's email-only notifications."),
    ("9", "Last-Minute Deals Section", "Dedicated section with countdown timers and urgency badges for departures within 7 days", "Creates urgency and surfaces deals that would otherwise be buried. Good conversion driver."),
    ("10", "Referral System", "Built-in referral codes with tracking, accessible from user profile", "Viral growth mechanism. Each user becomes a potential acquisition channel."),
]

for i, (num, feat, detail, why) in enumerate(strengths):
    r = 3 + i
    ws4.row_dimensions[r].height = 55
    bg = PatternFill("solid", fgColor=LIGHT_GREEN) if i % 2 == 0 else None
    style_cell(ws4, r, 1, num, alignment=wrap_center, fill=bg)
    style_cell(ws4, r, 2, feat, font=body_font_bold, fill=bg)
    style_cell(ws4, r, 3, detail, fill=bg)
    style_cell(ws4, r, 4, why, fill=bg)

ws4.freeze_panes = "A3"

# Save
output_path = "/sessions/vigilant-wonderful-wozniak/mnt/bookitfly/BookItFly_vs_Skyscanner_UX_Audit.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
